import os
import io
import zipfile
import time

from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook
from tqdm import tqdm

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload


# ============================================================
# CONFIG
# ============================================================

# Excel chứa danh sách MNV
EXCEL_FILE = "IDs.xlsx"

# Folder trên GOOGLE DRIVE chứa PDF
ROOT_FOLDER_ID = "1pHnflFoG5UiYefXgqm7rulsKfqGyxTRs"
#ROOT_FOLDER_ID = "1H0VLqYwdJKTsVDlZJbYhUo-M2gwsUCUg"

# Folder LOCAL để lưu PDF
OUTPUT_FOLDER = "result"

# ZIP LOCAL
OUTPUT_ZIP = "result.zip"

# Danh sách MNV không tìm thấy
NOT_FOUND_FILE = "not_found.txt"

# ------------------------------------------------------------
# DOWNLOAD SETTINGS
# ------------------------------------------------------------

# Số file download cùng lúc
MAX_WORKERS = 16

# Số lần retry nếu download lỗi
MAX_RETRIES = 5

# Chunk size: 4 MB
CHUNK_SIZE = 4 * 1024 * 1024

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly"
]


# ============================================================
# GOOGLE DRIVE LOGIN
# ============================================================

def get_drive_service():

    creds = None

    if os.path.exists("token.json"):

        print("Đang đọc token.json...")

        creds = Credentials.from_authorized_user_file(
            "token.json",
            SCOPES
        )

    if not creds or not creds.valid:

        if creds and creds.expired and creds.refresh_token:

            print("Token hết hạn, đang refresh...")

            creds.refresh(Request())

        else:

            print(
                "Đang mở trình duyệt để đăng nhập Google..."
            )

            flow = InstalledAppFlow.from_client_secrets_file(
                "credentials.json",
                SCOPES
            )

            creds = flow.run_local_server(
                port=0
            )

        with open(
            "token.json",
            "w",
            encoding="utf-8"
        ) as token:

            token.write(
                creds.to_json()
            )

    return build(
        "drive",
        "v3",
        credentials=creds,
        cache_discovery=False
    )


# ============================================================
# READ EXCEL
# ============================================================

def read_mnv():

    print("\n================================")
    print("ĐỌC EXCEL")
    print("================================")

    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(
            f"Không tìm thấy file Excel: {EXCEL_FILE}"
        )

    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True
    )

    sheet = workbook.active

    # A = STT
    # F = Mã nhân viên
    # Header ở dòng 1
    # Dữ liệu bắt đầu từ dòng 2
    STT_COLUMN = 1
    MNV_COLUMN = 6
    START_ROW = 2

    data = []

    for row in sheet.iter_rows(min_row=START_ROW):

        stt = row[STT_COLUMN - 1].value
        mnv = row[MNV_COLUMN - 1].value

        # Bỏ qua dòng trống
        if stt is None or mnv is None:
            continue

        stt = str(stt).strip()
        mnv = str(mnv).strip()

        if stt == "" or mnv == "":
            continue

        data.append({
            "stt": stt,
            "mnv": mnv
        })

    workbook.close()

    print(f"Đọc được {len(data)} nhân viên.")

    return data


# ============================================================
# GET ALL PDF FILES
# ============================================================

def get_all_pdfs(service):

    print("\n================================")
    print("QUÉT PDF TRÊN GOOGLE DRIVE")
    print("================================")

    print(
        f"Folder ID: {ROOT_FOLDER_ID}"
    )

    files = []

    page_token = None

    while True:

        response = service.files().list(

            q=(
                f"'{ROOT_FOLDER_ID}' in parents "
                "and trashed = false "
                "and mimeType = 'application/pdf'"
            ),

            includeItemsFromAllDrives=True,

            supportsAllDrives=True,

            pageSize=1000,

            fields=(
                "nextPageToken,"
                "files(id,name,mimeType)"
            ),

            pageToken=page_token

        ).execute()

        current_files = response.get(
            "files",
            []
        )

        files.extend(
            current_files
        )

        page_token = response.get(
            "nextPageToken"
        )

        if not page_token:
            break

    return files


# ============================================================
# MATCH FILES
# ============================================================

def create_file_map(data, files):

    print("\n================================")
    print("MATCH MNV")
    print("================================")

    result = {
        item["mnv"]: []
        for item in data
    }

    for file in tqdm(
        files,
        desc="Matching"
    ):

        filename = file["name"]

        for item in data:

            mnv = item["mnv"]

            if mnv in filename:

                result[mnv].append(
                    file
                )

    return result


# ============================================================
# DOWNLOAD ONE FILE
# ============================================================

def download_one_file(
    file_info,
    output_path
):
    """
    Mỗi worker tự tạo Google Drive service riêng.
    Không dùng chung service giữa các thread.
    """

    file_id = file_info["id"]

    original_name = file_info["name"]

    last_error = None

    for attempt in range(
        1,
        MAX_RETRIES + 1
    ):

        try:

            # ------------------------------------------------
            # Tạo service riêng cho thread
            # ------------------------------------------------

            service = get_drive_service()

            request = service.files().get_media(
                fileId=file_id
            )

            buffer = io.BytesIO()

            downloader = MediaIoBaseDownload(

                buffer,

                request,

                chunksize=CHUNK_SIZE

            )

            done = False

            while not done:

                _, done = downloader.next_chunk()

            # ------------------------------------------------
            # Write file
            # ------------------------------------------------

            with open(
                output_path,
                "wb"
            ) as output_file:

                output_file.write(
                    buffer.getvalue()
                )

            return {
                "success": True,
                "name": original_name,
                "path": output_path,
                "error": None
            }

        except Exception as e:

            last_error = e

            # ------------------------------------------------
            # Retry
            # ------------------------------------------------

            if attempt < MAX_RETRIES:

                wait_time = 2 ** attempt

                time.sleep(
                    wait_time
                )

            else:

                return {
                    "success": False,
                    "name": original_name,
                    "path": output_path,
                    "error": last_error
                }


# ============================================================
# PREPARE OUTPUT FOLDER
# ============================================================

def prepare_output_folder():

    print("\n================================")
    print("CHUẨN BỊ RESULT")
    print("================================")

    os.makedirs(
        OUTPUT_FOLDER,
        exist_ok=True
    )

    print(
        f"Output folder: {OUTPUT_FOLDER}"
    )


# ============================================================
# PREPARE DOWNLOAD TASKS
# ============================================================

def create_download_tasks(
    data,
    file_map
):

    tasks = []

    for item in data:

        stt = item["stt"]

        mnv = item["mnv"]

        matched_files = file_map[mnv]

        for file in matched_files:

            original_name = file["name"]

            output_name = (
                f"{stt}_{original_name}"
            )

            output_path = os.path.join(
                OUTPUT_FOLDER,
                output_name
            )

            tasks.append({
                "file": file,
                "output_path": output_path
            })

    return tasks


# ============================================================
# DOWNLOAD ALL FILES IN PARALLEL
# ============================================================

def download_files_parallel(
    tasks
):

    print("\n================================")
    print("FAST DOWNLOAD")
    print("================================")

    print(
        f"Concurrent downloads: {MAX_WORKERS}"
    )

    print(
        f"Total files: {len(tasks)}"
    )

    if not tasks:

        print(
            "Không có file để download."
        )

        return 0

    successful = 0

    failed = []

    # --------------------------------------------------------
    # Thread pool
    # --------------------------------------------------------

    with ThreadPoolExecutor(
        max_workers=MAX_WORKERS
    ) as executor:

        future_map = {}

        # ----------------------------------------------------
        # Submit all tasks
        # ----------------------------------------------------

        for task in tasks:

            future = executor.submit(

                download_one_file,

                task["file"],

                task["output_path"]

            )

            future_map[future] = task

        # ----------------------------------------------------
        # Process results
        # ----------------------------------------------------

        with tqdm(
            total=len(tasks),
            desc="Downloading",
            unit="file"
        ) as progress:

            for future in as_completed(
                future_map
            ):

                task = future_map[future]

                try:

                    result = future.result()

                    if result["success"]:

                        successful += 1

                    else:

                        failed.append(
                            result
                        )

                except Exception as e:

                    failed.append({

                        "success": False,

                        "name": task["file"]["name"],

                        "path": task["output_path"],

                        "error": e

                    })

                progress.update(1)

    # --------------------------------------------------------
    # Results
    # --------------------------------------------------------

    print("\n================================")
    print("DOWNLOAD RESULT")
    print("================================")

    print(
        f"Success : {successful}"
    )

    print(
        f"Failed  : {len(failed)}"
    )

    # --------------------------------------------------------
    # Failed files
    # --------------------------------------------------------

    if failed:

        print("\nCác file download lỗi:")

        for item in failed:

            print(
                f"- {item['name']}"
            )

            print(
                f"  Error: {item['error']}"
            )

    return successful


# ============================================================
# CREATE ZIP
# ============================================================

def create_zip():

    print("\n================================")
    print("TẠO ZIP")
    print("================================")

    if os.path.exists(
        OUTPUT_ZIP
    ):

        os.remove(
            OUTPUT_ZIP
        )

    files_to_zip = []

    if os.path.exists(
        OUTPUT_FOLDER
    ):

        for filename in os.listdir(
            OUTPUT_FOLDER
        ):

            filepath = os.path.join(
                OUTPUT_FOLDER,
                filename
            )

            if os.path.isfile(
                filepath
            ):

                files_to_zip.append(
                    filepath
                )

    if not files_to_zip:

        print(
            "Không có file để tạo ZIP."
        )

        return

    with zipfile.ZipFile(

        OUTPUT_ZIP,

        "w",

        compression=zipfile.ZIP_DEFLATED

    ) as zip_file:

        for filepath in tqdm(

            files_to_zip,

            desc="Zipping"

        ):

            filename = os.path.basename(
                filepath
            )

            zip_file.write(

                filepath,

                arcname=filename

            )

    print(
        f"\nĐã tạo: {OUTPUT_ZIP}"
    )


# ============================================================
# NOT FOUND REPORT
# ============================================================

def create_not_found_report(
    not_found
):

    with open(

        NOT_FOUND_FILE,

        "w",

        encoding="utf-8"

    ) as f:

        f.write(
            "DANH SACH MNV KHONG TIM THAY FILE\n"
        )

        f.write(
            "================================\n\n"
        )

        for item in not_found:

            f.write(

                f"STT: {item['stt']}    "
                f"MNV: {item['mnv']}\n"

            )

    print(
        f"\nĐã tạo: {NOT_FOUND_FILE}"
    )


# ============================================================
# MAIN
# ============================================================

def main():

    print()
    print("================================")
    print("GOOGLE DRIVE PDF FINDER")
    print("================================")

    # --------------------------------------------------------
    # 1. Excel
    # --------------------------------------------------------

    print("\n[1/6] Đọc Excel...")

    excel_data = read_mnv()

    print(
        f"Tìm thấy {len(excel_data)} MNV."
    )

    # --------------------------------------------------------
    # 2. Login
    # --------------------------------------------------------

    print("\n[2/6] Kết nối Google Drive...")

    service = get_drive_service()

    print(
        "Đăng nhập Google thành công."
    )

    # --------------------------------------------------------
    # 3. Prepare result folder
    # --------------------------------------------------------

    print("\n[3/6] Chuẩn bị folder result...")

    prepare_output_folder()

    # --------------------------------------------------------
    # 4. Scan PDF
    # --------------------------------------------------------

    print("\n[4/6] Quét PDF...")

    files = get_all_pdfs(
        service
    )

    print(
        f"Tìm thấy {len(files)} PDF."
    )

    # --------------------------------------------------------
    # 5. Match
    # --------------------------------------------------------

    print("\n[5/6] Match MNV...")

    file_map = create_file_map(

        excel_data,

        files

    )

    found = 0

    not_found = []

    total_matched_files = 0

    for item in excel_data:

        mnv = item["mnv"]

        matched = file_map[mnv]

        if matched:

            found += 1

            total_matched_files += len(
                matched
            )

        else:

            not_found.append(
                item
            )

    # --------------------------------------------------------
    # Statistics
    # --------------------------------------------------------

    print("\n================================")
    print("THỐNG KÊ")
    print("================================")

    print(
        f"Tổng MNV          : {len(excel_data)}"
    )

    print(
        f"MNV có file       : {found}"
    )

    print(
        f"MNV không có file : {len(not_found)}"
    )

    print(
        f"Tổng PDF match    : {total_matched_files}"
    )

    # --------------------------------------------------------
    # Not found
    # --------------------------------------------------------

    if not_found:

        create_not_found_report(
            not_found
        )

    else:

        print(
            "\nTất cả MNV đều có file match!"
        )

        if os.path.exists(
            NOT_FOUND_FILE
        ):

            os.remove(
                NOT_FOUND_FILE
            )

    # --------------------------------------------------------
    # Prepare tasks
    # --------------------------------------------------------

    tasks = create_download_tasks(

        excel_data,

        file_map

    )

    # --------------------------------------------------------
    # FAST DOWNLOAD
    # --------------------------------------------------------

    print(
        "\nBắt đầu download song song..."
    )

    total_downloaded = download_files_parallel(
        tasks
    )

    # --------------------------------------------------------
    # ZIP
    # --------------------------------------------------------

    print("\n[6/6] Tạo ZIP...")

    create_zip()

    # --------------------------------------------------------
    # Done
    # --------------------------------------------------------

    print("\n================================")
    print("HOÀN TẤT")
    print("================================")

    print(
        f"Downloaded: {total_downloaded}"
    )

    print(
        f"Folder: {OUTPUT_FOLDER}"
    )

    print(
        f"ZIP: {OUTPUT_ZIP}"
    )


# ============================================================
# START
# ============================================================

if __name__ == "__main__":

    main()
