from openpyxl import Workbook, load_workbook

from app.jobs.textnorm import cccd_key, find_column, is_numeric, locate_header_row, normalize_code, parse_date_value

# DS.xlsx đã được chuẩn hoá: STT luôn ở cột A (index 0), MNV luôn ở cột C (index 2).
# Không đọc thêm cột nào khác làm Họ và tên — các file thực tế đặt cột D khác nhau tuỳ nguồn
# (có khi là Họ và tên, có khi là Phòng/Vùng/Miền...), nên Họ và tên chỉ lấy từ báo cáo đào tạo.
DS_STT_COL = 0
DS_MNV_COL = 2

BAOCAO_MNV_CANDIDATES = ["Mã nhân viên", "Mã NV", "MNV"]
BAOCAO_CCCD_CANDIDATES = ["Số CMND/Hộ chiếu", "CCCD", "Số CMND"]
BAOCAO_HOTEN_CANDIDATES = ["Họ và tên", "Họ tên"]

# Các cột bổ sung dùng khi ghi đầy đủ 1 dòng mới vào sheet checklist (xem read_baocao_full_index).
# key ở đây phải khớp với field name dùng trong app/addhoso.
BAOCAO_EXTRA_FIELD_CANDIDATES: dict[str, list[str]] = {
    "job": ["Job"],
    "chuc_danh": ["Chức danh"],
    "gioi_tinh": ["Giới tính"],
    "ngay_sinh": ["Ngày sinh"],
    "noi_sinh": ["Nơi sinh"],
    "dia_chi_thuong_tru": ["Địa chỉ thường trú"],
    "dia_chi_hien_tai": ["Địa chỉ hiện tại"],
    "ngay_cap": ["Ngày cấp"],
    "noi_cap": ["Nơi cấp"],
    "dien_thoai": ["Điện thoại di động", "Điện thoại"],
    "ngay_vao_cong_ty": ["Ngày vào công ty"],
    # Header thật đôi khi có xuống dòng "Phòng/Vùng/\nMiền" -> sau chuẩn hoá thành có khoảng
    # trắng thừa "Phòng/Vùng/ Miền", nên liệt kê cả 2 biến thể.
    "phong_vung_mien": ["Phòng/Vùng/Miền", "Phòng/Vùng/ Miền"],
}


def _find_data_start_row(ws, max_scan: int = 20) -> int:
    """DS.xlsx có thể có nhiều dòng header phía trên (VD: 3 dòng) trước khi tới dữ liệu thật.
    Dò dòng đầu tiên có STT (cột A) là số để xác định dòng bắt đầu dữ liệu.
    """

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan), start=1):
        if is_numeric(row[DS_STT_COL].value):
            return row_idx
    return 2


def read_ds_list(path: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = workbook.active
        start_row = _find_data_start_row(ws)

        data = []
        for row in ws.iter_rows(min_row=start_row):
            stt_value = row[DS_STT_COL].value
            mnv_value = row[DS_MNV_COL].value

            mnv = normalize_code(mnv_value)
            if not mnv:
                continue

            data.append({"stt": normalize_code(stt_value), "mnv": mnv, "ho_ten": None})

        return data
    finally:
        workbook.close()


def read_cccd_list(path: str) -> list[dict]:
    """Đọc danh sách CCCD cần tìm — cùng định dạng chuẩn với DS.xlsx (STT cột A, mã cột C)."""

    raw = read_ds_list(path)
    return [{"stt": item["stt"], "cccd": item["mnv"], "ho_ten": item["ho_ten"]} for item in raw]


def _cell_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _read_baocao_rows(path: str) -> list[dict]:
    """Đọc file 'BÁO CÁO ĐÀO TẠO' -> list[dict] (mnv, cccd, ho_ten + các field ở
    BAOCAO_EXTRA_FIELD_CANDIDATES).

    Cột được dò theo TÊN header, không hardcode vị trí, vì file này đổi theo ngày
    và có thể lệch thứ tự cột. Ngoài mnv/cccd, mọi cột khác là best-effort — thiếu
    cột nào thì field đó trả về None, không raise lỗi (chỉ mnv/cccd là bắt buộc).
    """

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = workbook.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        if not rows:
            return []

        header_idx = locate_header_row(rows, [BAOCAO_MNV_CANDIDATES, BAOCAO_CCCD_CANDIDATES])
        header = rows[header_idx]
        data_rows = rows[header_idx + 1 :]

        idx_mnv = find_column(header, BAOCAO_MNV_CANDIDATES)
        idx_cccd = find_column(header, BAOCAO_CCCD_CANDIDATES)
        idx_hoten = find_column(header, BAOCAO_HOTEN_CANDIDATES)

        if idx_mnv is None or idx_cccd is None:
            raise ValueError(
                "Không tìm thấy cột 'Mã nhân viên' hoặc 'Số CMND/Hộ chiếu' trong file báo cáo đào tạo."
            )

        extra_idx = {key: find_column(header, candidates) for key, candidates in BAOCAO_EXTRA_FIELD_CANDIDATES.items()}

        result = []
        for row in data_rows:
            if idx_mnv >= len(row):
                continue

            mnv = normalize_code(row[idx_mnv])
            if not mnv:
                continue

            cccd = normalize_code(row[idx_cccd]) if idx_cccd < len(row) else ""
            ho_ten = _cell_text(row[idx_hoten]) if idx_hoten is not None and idx_hoten < len(row) else None

            entry = {"mnv": mnv, "cccd": cccd, "ho_ten": ho_ten}
            for key, idx in extra_idx.items():
                entry[key] = row[idx] if idx is not None and idx < len(row) else None
                if key in ("ngay_sinh", "ngay_cap", "ngay_vao_cong_ty"):
                    # Có file lưu 3 cột này dạng Date thật, có file lại lưu dạng text
                    # (VD: "01/04/2026") -> chuẩn hoá về datetime cho cả 2 trường hợp.
                    entry[key] = parse_date_value(entry[key])
                else:
                    entry[key] = _cell_text(entry[key])

            result.append(entry)

        return result
    finally:
        workbook.close()


def read_baocao_index(path: str) -> dict[str, dict]:
    """dict[MNV] = {"cccd": str, "ho_ten": str|None} — dùng khi đã biết MNV."""

    return {r["mnv"]: {"cccd": r["cccd"], "ho_ten": r["ho_ten"]} for r in _read_baocao_rows(path)}


def read_baocao_cccd_index(path: str) -> dict[str, dict]:
    """dict[cccd_key(CCCD)] = {"mnv": str, "ho_ten": str|None} — chiều ngược lại read_baocao_index()."""

    result: dict[str, dict] = {}
    for r in _read_baocao_rows(path):
        if r["cccd"]:
            result[cccd_key(r["cccd"])] = {"mnv": r["mnv"], "ho_ten": r["ho_ten"]}
    return result


def read_baocao_full_index(path: str) -> dict[str, dict]:
    """dict[cccd_key(CCCD)] = toàn bộ thông tin nhân viên (mnv, ho_ten + các field cá nhân khác)
    lấy từ báo cáo đào tạo — dùng để fill đầy đủ 1 dòng mới vào sheet checklist.
    """

    result: dict[str, dict] = {}
    for r in _read_baocao_rows(path):
        if r["cccd"]:
            result[cccd_key(r["cccd"])] = r
    return result


NOT_FOUND_COLUMNS = [
    ("stt", "STT"),
    ("mnv", "MNV"),
    ("ho_ten", "Họ và tên"),
    ("cccd", "CCCD"),
    ("phong_tuyen_dung", "Phòng tuyển dụng"),
    ("da_nhan_ho_so", "Đã nhận hồ sơ"),
    ("nguon", "Nguồn"),
    ("ngay_viet_ho_so", "Ngày viết hồ sơ"),
    ("co_ma_checklist", "Có mã (checklist)"),
    ("ket_luan", "Kết luận"),
]


def write_not_found_excel(rows: list[dict], output_path: str) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Not found"

    ws.append([label for _, label in NOT_FOUND_COLUMNS])
    for row in rows:
        ws.append([row.get(key, "") for key, _ in NOT_FOUND_COLUMNS])

    workbook.save(output_path)
